"""
Excel导出智能体
负责将测试用例导出为Excel格式
"""
import os
import logging
from typing import Any, Dict, List, Optional
from datetime import datetime
from pathlib import Path

from autogen_core import MessageContext, type_subscription
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.agents.base import BaseAgent
from app.core.types import AgentTypes, TopicTypes, AGENT_NAMES
from app.database.models.test_case import TestCase
from app.core.messages.test_case import ExcelExportRequest
from app.database.connection import db_manager
from sqlalchemy import select

logger = logging.getLogger(__name__)


@type_subscription(topic_type=TopicTypes.EXCEL_EXPORTER.value)
class ExcelExporterAgent(BaseAgent):
    """Excel导出智能体，负责将测试用例导出为Excel格式"""

    def __init__(self, model_client_instance=None, **kwargs):
        """初始化Excel导出智能体"""
        super().__init__(
            agent_id=AgentTypes.EXCEL_EXPORTER.value,
            agent_name=AGENT_NAMES[AgentTypes.EXCEL_EXPORTER.value],
            model_client_instance=model_client_instance,
            **kwargs
        )

        self.export_records: Dict[str, Dict[str, Any]] = {}
        
        # 确保导出目录存在
        self.export_dir = Path("exports/excel")
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"Excel导出智能体初始化完成: {self.agent_name}")

    async def handle_export_request(self, message: ExcelExportRequest, ctx: MessageContext) -> None:
        """处理Excel导出请求"""
        try:
            export_id = f"export_{message.session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            await self.send_response(f"🚀 开始Excel导出: {export_id}")
            
            # 记录导出开始
            self.export_records[export_id] = {
                "export_id": export_id,
                "session_id": message.session_id,
                "start_time": datetime.now(),
                "status": "processing"
            }
            
            # 获取测试用例数据
            test_cases = await self._get_test_cases(message)
            
            if not test_cases:
                await self.send_response("⚠️ 没有找到要导出的测试用例", is_final=True)
                return
            
            await self.send_response(f"📊 找到 {len(test_cases)} 个测试用例，开始生成Excel文件")
            
            # 生成Excel文件
            file_path = await self._generate_excel_file(export_id, test_cases, message)
            
            # 更新导出记录
            self.export_records[export_id].update({
                "end_time": datetime.now(),
                "status": "completed",
                "file_path": str(file_path),
                "test_case_count": len(test_cases)
            })
            
            await self.send_response(
                f"✅ Excel导出完成: {file_path.name}",
                is_final=True,
                result={
                    "export_id": export_id,
                    "file_path": str(file_path),
                    "test_case_count": len(test_cases)
                }
            )
            
        except Exception as e:
            logger.error(f"Excel导出失败: {str(e)}")
            if export_id in self.export_records:
                self.export_records[export_id].update({
                    "end_time": datetime.now(),
                    "status": "failed",
                    "error": str(e)
                })
            await self.send_response(f"❌ Excel导出失败: {str(e)}", is_final=True)

    async def _get_test_cases(self, request: ExcelExportRequest) -> List[TestCase]:
        """获取要导出的测试用例"""
        try:
            async with db_manager.get_session() as session:
                if request.test_case_ids:
                    # 导出指定的测试用例
                    stmt = select(TestCase).where(TestCase.id.in_(request.test_case_ids))
                    result = await session.execute(stmt)
                    return result.scalars().all()
                else:
                    # 导出会话中的所有测试用例
                    stmt = select(TestCase).where(TestCase.session_id == request.session_id)
                    result = await session.execute(stmt)
                    return result.scalars().all()

        except Exception as e:
            logger.error(f"获取测试用例失败: {str(e)}")
            return []

    async def _generate_excel_file(self, export_id: str, test_cases: List[TestCase], request: ExcelExportRequest) -> Path:
        """生成Excel文件"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "测试用例"
            
            # 设置列标题
            headers = [
                "序号", "测试用例ID", "测试用例名称", "测试类型", "测试级别", 
                "优先级", "前置条件", "测试步骤", "预期结果", "标签", 
                "分类", "状态", "创建时间", "更新时间"
            ]
            
            # 写入标题行
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            
            # 写入数据行
            for row, test_case in enumerate(test_cases, 2):
                data = [
                    row - 1,  # 序号
                    test_case.id,
                    test_case.title,
                    test_case.test_type,
                    test_case.test_level,
                    test_case.priority,
                    test_case.preconditions or "",
                    self._format_test_steps(test_case.test_steps),
                    test_case.expected_result or "",
                    ", ".join(test_case.tags) if test_case.tags else "",
                    test_case.category or "",
                    test_case.status,
                    test_case.created_at.strftime("%Y-%m-%d %H:%M:%S") if test_case.created_at else "",
                    test_case.updated_at.strftime("%Y-%m-%d %H:%M:%S") if test_case.updated_at else ""
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
            
            # 调整列宽
            column_widths = [8, 15, 30, 12, 12, 10, 25, 40, 25, 20, 15, 10, 20, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # 保存文件
            filename = f"测试用例导出_{export_id}.xlsx"
            file_path = self.export_dir / filename
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成Excel文件失败: {str(e)}")
            raise

    def _format_test_steps(self, test_steps: List[Dict[str, Any]]) -> str:
        """格式化测试步骤"""
        if not test_steps:
            return ""
        
        formatted_steps = []
        for i, step in enumerate(test_steps, 1):
            step_text = f"{i}. {step.get('action', '')}"
            if step.get('expected'):
                step_text += f"\n   预期: {step['expected']}"
            formatted_steps.append(step_text)
        
        return "\n".join(formatted_steps)

    def get_export_status(self, export_id: str) -> Optional[Dict[str, Any]]:
        """获取导出状态"""
        return self.export_records.get(export_id)

    def list_exports(self) -> List[Dict[str, Any]]:
        """列出所有导出记录"""
        return list(self.export_records.values())
